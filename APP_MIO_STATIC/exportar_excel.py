"""
Exporta datos de APP MIO a Excel (.xlsx).

Uso:
  1. En la app: Configuracion > Exportar datos (JSON)
  2. pip install openpyxl
  3. python exportar_excel.py APP_MIO_datos_2025-01-01.json
"""
import json
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def fmt_date(iso):
    if not iso:
        return ''
    try:
        return datetime.fromisoformat(iso[:10]).strftime('%d/%m/%Y')
    except Exception:
        return iso[:10]

def main():
    json_file = sys.argv[1] if len(sys.argv) > 1 else 'datos.json'
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f'Error: no se encontro el archivo "{json_file}"')
        print('Uso: python exportar_excel.py APP_MIO_datos.json')
        sys.exit(1)

    people   = data.get('people', [])
    payrolls = data.get('payrolls', [])

    wb = openpyxl.Workbook()

    # ── Estilos ──────────────────────────────────────────────
    hdr_fill  = PatternFill('solid', fgColor='6366F1')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin', color='D1D5DB')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill('solid', fgColor='F0F0FF')
    center    = Alignment(horizontal='center', vertical='center')
    left      = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def style_header(ws, headers, col_widths=None):
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = hdr_align
            c.border = border
        if col_widths:
            for col, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(1, col).column_letter].width = w
        ws.row_dimensions[1].height = 28

    def style_row(ws, row_num, values, aligns=None):
        fill = alt_fill if row_num % 2 == 0 else None
        for col, val in enumerate(values, 1):
            c = ws.cell(row_num, col, val)
            c.border = border
            if fill:
                c.fill = fill
            c.alignment = (aligns[col-1] if aligns and col-1 < len(aligns) else left)

    # ── Hoja: Personas ───────────────────────────────────────
    ws_p = wb.active
    ws_p.title = 'Personas'
    ws_p.freeze_panes = 'A2'

    p_headers   = ['Nombre completo', 'DNI / Cedula', 'Pasaporte',
                   'Bancos', 'IBANs', 'N. cuenta', 'Wallets (red + direccion)', 'Registrado']
    p_widths    = [30, 16, 16, 22, 30, 22, 40, 14]
    style_header(ws_p, p_headers, p_widths)

    for r, p in enumerate(people, 2):
        banks  = p.get('bankAccounts', [])
        wallets = p.get('wallets', [])
        style_row(ws_p, r, [
            p.get('fullName', ''),
            p.get('dni', ''),
            p.get('passport', ''),
            ', '.join(b.get('bankName','') for b in banks if b.get('bankName')),
            ', '.join(b.get('iban','')     for b in banks if b.get('iban')),
            ', '.join(b.get('accountNumber','') for b in banks if b.get('accountNumber')),
            '; '.join(f"{w.get('network','')} {w.get('address','')}".strip() for w in wallets),
            fmt_date(p.get('createdAt','')),
        ])

    # ── Hoja: Nominas ────────────────────────────────────────
    ws_n = wb.create_sheet('Nominas')
    ws_n.freeze_panes = 'A2'

    n_headers = ['Persona', 'DNI', 'Monto ($)', 'Fecha vigencia', 'Notas', 'Registrado']
    n_widths  = [30, 16, 14, 16, 40, 14]
    style_header(ws_n, n_headers, n_widths)

    p_map = {p['id']: p for p in people}
    for r, pay in enumerate(sorted(payrolls, key=lambda x: x.get('effectiveDate',''), reverse=True), 2):
        p = p_map.get(pay.get('personId'), {})
        style_row(ws_n, r, [
            p.get('fullName', '(eliminada)'),
            p.get('dni', ''),
            float(pay.get('amount', 0)),
            pay.get('effectiveDate', ''),
            pay.get('notes', ''),
            fmt_date(pay.get('createdAt','')),
        ], aligns=[left, center, center, center, left, center])
        ws_n.cell(r, 3).number_format = '#,##0.00'

    # ── Guardar ──────────────────────────────────────────────
    out = f'APP_MIO_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(out)
    print(f'Exportado: {out}')
    print(f'  {len(people)} personas  |  {len(payrolls)} nominas')

if __name__ == '__main__':
    main()
